

from colorama import Fore
from fpdf import FPDF
from openpyxl import load_workbook
import excel2img

wb = load_workbook('testfile.xlsx', data_only = True)
ws1 = wb['Sheet1']

achieved_amount = str(int(ws1['G8'].value*100)) + '%'
achieved_amount_text = Fore.RED + achieved_amount
total_central = str(int(ws1['I16'].value*100))+'%'

img = excel2img.export_img('testfile.xlsx','testimage.png',"",'Sheet1!E12:I20')

print(achieved_amount_text)


name = 'aaa'

def create_title(pdf):
    pdf.set_font('Arial', '', 24)
    pdf.ln(20)
    pdf.write(5,f"ABC Analytics Report")
    pdf.ln(10)
    pdf.write(4,'2023/1/1')
    pdf.ln(20)

'''First Page'''

pdf = FPDF()
pdf.add_page()
create_title(pdf)
pdf.set_font_size(12)
pdf.cell(0,10,f'Today the amount is {total_central},achieved amount is /n {achieved_amount_text}')
pdf.image('testimage.png',20.60,100,150)
pdf.cell(10,400,f'xxxx')
pdf.output('tutorial.pdf')


# pdf = FPDF()
# pdf.add_page()
# create_title(pdf)
# pdf.cell(40,10,f'Hello my name is {name}!')
# pdf.image('save_fig.png',5,100, 80)
# pdf.image('save_fig.png',105,100, 80)
# pdf.add_page()
# pdf.image('save_fig.png',5,80, 80)
# pdf.image('save_fig.png',105,80, 80)
# pdf.cell(40,205,f'Hello my name is {name}!')
# pdf.output('tutorial.pdf', 'F')
